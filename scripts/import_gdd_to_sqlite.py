from __future__ import annotations

import argparse
import os
import sqlite3
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


DEFAULT_SHEET = "头季"
DEFAULT_TABLE = "gdd_stages"
STAGE_COLUMNS = {
    "三叶一心",
    "返青",
    "分蘖期",
    "有效分蘖终止期",
    "拔节期",
    "幼穗分化1期",
    "幼穗分化2期",
    "幼穗分化4期",
    "孕穗期",
    "破口期",
    "始穗期",
    "抽穗期",
    "齐穗期",
    "成熟期",
}


def _resolve_path(value: str, *, must_exist: bool = False) -> Path:
    path = Path(value).expanduser()
    if must_exist and not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")
    return path


def _load_excel(path: Path, sheet_name: str) -> Tuple[List[str], List[List[object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Excel 缺少工作表: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    headers = [
        str(cell).strip() if cell is not None else ""
        for cell in header_row
    ]
    data_rows: List[List[object]] = []
    for row in rows:
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        data_rows.append(list(row))
    workbook.close()
    return headers, data_rows


def _infer_column_types(headers: List[str]) -> Dict[str, str]:
    types: Dict[str, str] = {}
    for header in headers:
        if not header:
            continue
        if header in STAGE_COLUMNS:
            types[header] = "REAL"
        else:
            types[header] = "TEXT"
    return types


def _build_schema(headers: List[str], table: str) -> str:
    column_types = _infer_column_types(headers)
    columns = []
    for header in headers:
        if not header:
            continue
        col_type = column_types.get(header, "TEXT")
        columns.append(f'"{header}" {col_type}')
    columns_sql = ", ".join(["id INTEGER PRIMARY KEY AUTOINCREMENT"] + columns)
    return f'CREATE TABLE IF NOT EXISTS "{table}" ({columns_sql})'


def _insert_rows(
    conn: sqlite3.Connection,
    table: str,
    headers: List[str],
    rows: List[List[object]],
) -> None:
    valid_headers = [h for h in headers if h]
    if not valid_headers or not rows:
        return
    placeholders = ", ".join(["?"] * len(valid_headers))
    quoted_cols = ", ".join(f'"{h}"' for h in valid_headers)
    sql = f'INSERT INTO "{table}" ({quoted_cols}) VALUES ({placeholders})'
    payloads = []
    for row in rows:
        values = [row[idx] if idx < len(row) else None for idx, _ in enumerate(headers)]
        filtered = [
            values[idx] for idx, header in enumerate(headers) if header
        ]
        payloads.append(filtered)
    conn.executemany(sql, payloads)


def _create_indexes(conn: sqlite3.Connection, table: str, headers: List[str]) -> None:
    for column in ("审定地区", "稻作类型", "亚种", "熟制", "标准品种", "品种"):
        if column not in headers:
            continue
        conn.execute(
            f'CREATE INDEX IF NOT EXISTS "idx_{table}_{column}" '
            f'ON "{table}" ("{column}")'
        )


def import_excel_to_sqlite(
    excel_path: Path,
    db_path: Path,
    *,
    sheet_name: str,
    table: str,
    replace: bool,
) -> None:
    headers, rows = _load_excel(excel_path, sheet_name)
    if not headers:
        raise ValueError("Excel 表头为空，无法导入。")
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        if replace:
            conn.execute(f'DROP TABLE IF EXISTS "{table}"')
        conn.execute(_build_schema(headers, table))
        conn.execute(f'DELETE FROM "{table}"')
        _insert_rows(conn, table, headers, rows)
        _create_indexes(conn, table, headers)
        conn.commit()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="将 GDD Excel 导入 SQLite。"
    )
    parser.add_argument(
        "--excel",
        default=os.getenv("GROWTH_STAGE_GDD_PATH"),
        required=False,
        help="GDD Excel 路径（可用 GROWTH_STAGE_GDD_PATH）。",
    )
    parser.add_argument(
        "--db",
        default=os.getenv("GROWTH_STAGE_DB_PATH"),
        required=False,
        help="SQLite 路径（可用 GROWTH_STAGE_DB_PATH）。",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help="Excel 工作表名称（默认 头季）。",
    )
    parser.add_argument(
        "--table",
        default=DEFAULT_TABLE,
        help="SQLite 表名（默认 gdd_stages）。",
    )
    parser.add_argument(
        "--append",
        action="store_true",
        help="追加模式（不覆盖表结构）。",
    )
    args = parser.parse_args()

    if not args.excel:
        raise SystemExit("缺少 Excel 路径，请使用 --excel 或设置 GROWTH_STAGE_GDD_PATH。")
    excel_path = _resolve_path(args.excel, must_exist=True)
    if args.db:
        db_path = _resolve_path(args.db)
    else:
        db_path = (
            Path(__file__).resolve().parents[1]
            / "resources"
            / "gdd.sqlite3"
        )
    replace = not args.append
    import_excel_to_sqlite(
        excel_path,
        db_path,
        sheet_name=args.sheet,
        table=args.table,
        replace=replace,
    )
    mode = "追加" if args.append else "覆盖"
    print(f"已导入 GDD 数据 -> {db_path} ({mode}模式)")


if __name__ == "__main__":
    main()
